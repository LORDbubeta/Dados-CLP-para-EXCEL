from pyModbusTCP.client import ModbusClient
import time
from openpyxl import Workbook, load_workbook
import os

usuario = os.getlogin()

caminho = 'C:\\Users\\' + usuario + '\\Desktop\\Coleta de dados\\dados.xlsx'

#defino meu excel
planilha = Workbook()

#verifico se planilha existe
if not os.path.exists(caminho):

     planilha.create_sheet("dados", 0) 
     planilha.save(caminho)

else:

    print("planilha existente")

#defino meu clp e crio o objeto
ip = "192.168.1.5"
porta = 502
id = 1
t_out = 10.0

clp_delta = ModbusClient(host= ip, port= porta, unit_id= id, timeout= t_out, 
                         auto_open= True, auto_close= False) # type: ignore


def conectar_clp(clp):

    #conectar no clp e manter a conexão
    if clp_delta.is_open:

        print("conexão aberta")
    
    else:

        if clp_delta.open():

            print(f"conexão estabelecida com sucesso")
        
        else:
            
            print("conexão não estabelecida", clp_delta.last_error)

conectar_clp(clp = clp_delta)

def ler_registros():

    #ler valores do clp
    #armazenar na lista
    valores = []
    A = 20006
    registro_aux = 0

    try:       
        for aux in range(0,40,2):
            registros = clp_delta.read_holding_registers(A, 2)
            print(registros)

            if registros[1] > 0:
                registro_aux = registros[1]
                registro_aux_2 = registros[0] - registro_aux
                valores.append(registro_aux_2)  # Pegando o valor convertido para negativo

            else:
                valores.append(registros[0])  # Pegando o primeiro valor da lista

            A = A + 2
            print(valores)

    except Exception as e:
        print("Erro ao ler registros: {e}")

    if any(v is None for v in valores):
        print("Há valores indefinidos na lista.")
    else:
        print("Não há valores vazios na lista.")
        enviar_excel(valor = valores)   


def enviar_excel(valor):

    try:

            #abrir excel
            planilha = load_workbook(caminho)
            Plan = planilha.active
    
            #enviar valores para o excel pegando da lista
            #usar um for para colocar os valores no excel
            for i, v in enumerate(valor, start=1):
                Plan.cell(row=i, column=1, value=v)

            #salvar e fechar excel
            planilha.save(caminho)

            print("dados escrito com sucesso") 

    except:

        print("ocorreu algum erro na escrita")
    

while True:

    ler_registros()
    time.sleep(30)
